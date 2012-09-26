"""
Module contains tools for processing files into DataFrames or other objects
"""
from StringIO import StringIO
import re
from itertools import izip
from urlparse import urlparse
import csv

try:
    next
except NameError:  # pragma: no cover
    # Python < 2.6
    def next(x):
        return x.next()

import numpy as np

from pandas.core.index import Index, MultiIndex
from pandas.core.frame import DataFrame
import datetime
import pandas.core.common as com
from pandas.util import py3compat
from pandas.io.date_converters import generic_parser

from pandas.util.decorators import Appender

import pandas.lib as lib
import pandas._parser as _parser

class DateConversionError(Exception):
    pass

_parser_params = """Also supports optionally iterating or breaking of the file
into chunks.

Parameters
----------
filepath_or_buffer : string or file handle / StringIO. The string could be
    a URL. Valid URL schemes include http, ftp, and file. For file URLs, a host
    is expected. For instance, a local file could be
    file ://localhost/path/to/table.csv
%s
dialect : string or csv.Dialect instance, default None
    If None defaults to Excel dialect. Ignored if sep longer than 1 char
    See csv.Dialect documentation for more details
header : int, default 0
    Row to use for the column labels of the parsed DataFrame
skiprows : list-like or integer
    Row numbers to skip (0-indexed) or number of rows to skip (int)
    at the start of the file
skip_footer : int, default 0
    Lines at bottom of file to skip. If >0 then indicates the row to start
    skipping. If <0 then skips the specified number of rows from the end.
index_col : int or sequence, default None
    Column to use as the row labels of the DataFrame. If a sequence is
    given, a MultiIndex is used.
names : array-like
    List of column names
na_values : list-like or dict, default None
    Additional strings to recognize as NA/NaN. If dict passed, specific
    per-column NA values
keep_default_na : bool, default True
    If na_values are specified and keep_default_na is False the default NaN
    values are overridden, otherwise they're appended to
parse_dates : boolean, list of ints or names, list of lists, or dict
    If True -> try parsing the index.
    If [1, 2, 3] -> try parsing columns 1, 2, 3 each as a separate date column.
    If [[1, 3]] -> combine columns 1 and 3 and parse as a single date column.
    {'foo' : [1, 3]} -> parse columns 1, 3 as date and call result 'foo'
keep_date_col : boolean, default False
    If True and parse_dates specifies combining multiple columns then
    keep the original columns.
date_parser : function
    Function to use for converting dates to strings. Defaults to
    dateutil.parser
dayfirst : boolean, default False
    DD/MM format dates, international and European format
thousands : str, default None
    Thousands separator
comment : str, default None
    Indicates remainder of line should not be parsed
    Does not support line commenting (will return empty line)
nrows : int, default None
    Number of rows of file to read. Useful for reading pieces of large files
iterator : boolean, default False
    Return TextParser object
chunksize : int, default None
    Return TextParser object for iteration
converters : dict. optional
    Dict of functions for converting values in certain columns. Keys can either
    be integers or column labels
verbose : boolean, default False
    Indicate number of NA values placed in non-numeric columns
delimiter : string, default None
    Alternative argument name for sep. Regular expressions are accepted.
encoding : string, default None
    Encoding to use for UTF when reading/writing (ex. 'utf-8')
squeeze : boolean, default False
    If the parsed data only contains one column then return a Series

Returns
-------
result : DataFrame or TextParser
"""

_csv_sep = """sep : string, default ','
    Delimiter to use. If sep is None, will try to automatically determine
    this. Regular expressions are accepted.
"""

_table_sep = """sep : string, default \\t (tab-stop)
    Delimiter to use. Regular expressions are accepted."""

_read_csv_doc = """
Read CSV (comma-separated) file into DataFrame

%s
""" % (_parser_params % _csv_sep)

_read_table_doc = """
Read general delimited file into DataFrame

%s
""" % (_parser_params % _table_sep)

_fwf_widths = """\
colspecs : a list of pairs (tuples), giving the extents
    of the fixed-width fields of each line as half-open internals
    (i.e.,  [from, to[  ).
widths : a list of field widths, which can be used instead of
    'colspecs' if the intervals are contiguous.
"""

_read_fwf_doc = """
Read a table of fixed-width formatted lines into DataFrame

%s

Also, 'delimiter' is used to specify the filler character of the
fields if it is not spaces (e.g., '~').
""" % (_parser_params % _fwf_widths)


def _is_url(url):
    """
    Very naive check to see if url is an http(s), ftp, or file location.
    """
    parsed_url = urlparse(url)
    if parsed_url.scheme in ['http','file', 'ftp', 'https']:
        return True
    else:
        return False

def _read(filepath_or_buffer, kwds):
    "Generic reader of line files."
    encoding = kwds.get('encoding', None)

    if isinstance(filepath_or_buffer, str) and _is_url(filepath_or_buffer):
        from urllib2 import urlopen
        filepath_or_buffer = urlopen(filepath_or_buffer)
        if py3compat.PY3:  # pragma: no cover
            if encoding:
                errors = 'strict'
            else:
                errors = 'replace'
                encoding = 'utf-8'
            bytes = filepath_or_buffer.read()
            filepath_or_buffer = StringIO(bytes.decode(encoding, errors))

    if hasattr(filepath_or_buffer, 'read'):
        f = filepath_or_buffer
    else:
        try:
            # universal newline mode
            f = com._get_handle(filepath_or_buffer, 'U', encoding=encoding)
        except Exception: # pragma: no cover
            f = com._get_handle(filepath_or_buffer, 'r', encoding=encoding)

    if kwds.get('date_parser', None) is not None:
        if isinstance(kwds['parse_dates'], bool):
            kwds['parse_dates'] = True

    # Extract some of the arguments (pass chunksize on).
    kwds.pop('filepath_or_buffer')
    iterator = kwds.pop('iterator')
    nrows = kwds.pop('nrows')
    chunksize = kwds.get('chunksize', None)

    # Create the parser.
    parser = TextFileReader(f, **kwds)

    if nrows is not None:
        return parser.read(nrows)
    elif chunksize or iterator:
        return parser

    return parser.read()

@Appender(_read_csv_doc)
def read_csv(filepath_or_buffer,
             sep=',',
             engine='python',
             dialect=None,
             header=0,
             index_col=None,
             names=None,
             skiprows=None,
             na_values=None,
             na_filter=True,
             keep_default_na=True,
             thousands=None,
             comment=None,
             parse_dates=False,
             keep_date_col=False,
             dayfirst=False,
             date_parser=None,
             nrows=None,
             iterator=False,
             chunksize=None,
             skip_footer=0,
             converters=None,
             verbose=False,
             delimiter=None,
             encoding=None,
             squeeze=False):
    kwds = dict(filepath_or_buffer=filepath_or_buffer,
                sep=sep, dialect=dialect,
                header=header, index_col=index_col,
                names=names, skiprows=skiprows,
                na_values=na_values,
                na_filter=na_filter,
                keep_default_na=keep_default_na,
                thousands=thousands,
                comment=comment, parse_dates=parse_dates,
                keep_date_col=keep_date_col,
                dayfirst=dayfirst, date_parser=date_parser,
                nrows=nrows, iterator=iterator,
                chunksize=chunksize, skip_footer=skip_footer,
                converters=converters, verbose=verbose,
                delimiter=delimiter, encoding=encoding,
                squeeze=squeeze,
                engine=engine)

    # Alias sep -> delimiter.
    sep = kwds.pop('sep')
    if kwds.get('delimiter', None) is None:
        kwds['delimiter'] = sep

    return _read(filepath_or_buffer, kwds)

@Appender(_read_table_doc)
def read_table(filepath_or_buffer,
               sep='\t',
               engine='python',
               dialect=None,
               header=0,
               index_col=None,
               names=None,
               skiprows=None,
               na_values=None,
               na_filter=True,
               keep_default_na=True,
               thousands=None,
               comment=None,
               parse_dates=False,
               keep_date_col=False,
               dayfirst=False,
               date_parser=None,
               nrows=None,
               iterator=False,
               chunksize=None,
               skip_footer=0,
               converters=None,
               verbose=False,
               delimiter=None,
               encoding=None,
               squeeze=False):
    kwds = dict(filepath_or_buffer=filepath_or_buffer,
                sep=sep, dialect=dialect,
                header=header, index_col=index_col,
                names=names, skiprows=skiprows,
                na_values=na_values,
                na_filter=na_filter,
                keep_default_na=keep_default_na,
                thousands=thousands,
                comment=comment, parse_dates=parse_dates,
                keep_date_col=keep_date_col,
                dayfirst=dayfirst, date_parser=date_parser,
                nrows=nrows, iterator=iterator,
                chunksize=chunksize, skip_footer=skip_footer,
                converters=converters, verbose=verbose,
                delimiter=delimiter, encoding=encoding,
                squeeze=squeeze)

    # Alias sep -> delimiter.
    sep = kwds.pop('sep')
    if kwds.get('delimiter', None) is None:
        kwds['delimiter'] = sep

    # Override as default encoding.
    kwds['encoding'] = None

    return _read(filepath_or_buffer, kwds)

@Appender(_read_fwf_doc)
def read_fwf(filepath_or_buffer,
             colspecs=None,
             widths=None,
             header=0,
             index_col=None,
             names=None,
             skiprows=None,
             na_values=None,
             na_filter=True,
             keep_default_na=True,
             thousands=None,
             comment=None,
             parse_dates=False,
             keep_date_col=False,
             dayfirst=False,
             date_parser=None,
             nrows=None,
             iterator=False,
             chunksize=None,
             skip_footer=0,
             converters=None,
             delimiter=None,
             verbose=False,
             encoding=None,
             squeeze=False):
    kwds = dict(filepath_or_buffer=filepath_or_buffer,
                colspecs=colspecs, widths=widths,
                header=header, index_col=index_col,
                names=names, skiprows=skiprows,
                na_values=na_values,
                na_filter=na_filter,
                keep_default_na=keep_default_na,
                thousands=thousands,
                comment=comment, parse_dates=parse_dates,
                keep_date_col=keep_date_col,
                dayfirst=dayfirst, date_parser=date_parser,
                nrows=nrows, iterator=iterator,
                chunksize=chunksize, skip_footer=skip_footer,
                converters=converters, verbose=verbose,
                delimiter=delimiter, encoding=encoding,
                squeeze=squeeze)

    # Check input arguments.
    colspecs = kwds.get('colspecs', None)
    widths = kwds.pop('widths', None)
    if bool(colspecs is None) == bool(widths is None):
        raise ValueError("You must specify only one of 'widths' and "
                         "'colspecs'")

    # Compute 'colspec' from 'widths', if specified.
    if widths is not None:
        colspecs, col = [], 0
        for w in widths:
            colspecs.append( (col, col+w) )
            col += w
        kwds['colspecs'] = colspecs

    kwds['thousands'] = thousands
    kwds['engine'] = 'python-fwf'

    return _read(filepath_or_buffer, kwds)

def read_clipboard(**kwargs):  # pragma: no cover
    """
    Read text from clipboard and pass to read_table. See read_table for the
    full argument list

    Returns
    -------
    parsed : DataFrame
    """
    from pandas.util.clipboard import clipboard_get
    text = clipboard_get()
    return read_table(StringIO(text), **kwargs)

def to_clipboard(obj): # pragma: no cover
    """
    Attempt to write text representation of object to the system clipboard

    Notes
    -----
    Requirements for your platform
      - Linux: xsel command line tool
      - Windows: Python win32 extensions
      - OS X:
    """
    from pandas.util.clipboard import clipboard_set
    clipboard_set(str(obj))


# common NA values
# no longer excluding inf representations
# '1.#INF','-1.#INF', '1.#INF000000',
_NA_VALUES = set(['-1.#IND', '1.#QNAN', '1.#IND', '-1.#QNAN',
                 '#N/A N/A', 'NA', '#NA', 'NULL', 'NaN',
                 'nan', ''])


class TextFileReader(object):
    """

    Passed dialect overrides any of the related parser options

    """

    def __init__(self, f,
                 engine='python', delimiter=None, dialect=None,
                 names=None, header=0, index_col=None,
                 thousands=None,
                 comment=None,
                 parse_dates=False, keep_date_col=False,
                 date_parser=None, dayfirst=False,
                 chunksize=None,
                 skiprows=None, skip_footer=0,
                 converters=None,
                 verbose=False,
                 encoding=None,
                 na_filter=True, na_values=None, keep_default_na=True,
                 warn_bad_lines=True, error_bad_lines=True,
                 doublequote=True,
                 escapechar=None,
                 quotechar='"',
                 quoting=csv.QUOTE_MINIMAL,
                 skipinitialspace=False,
                 squeeze=False,
                 as_recarray=False,
                 factorize=True,
                 **kwds):

        # Tokenization options
        self.delimiter = delimiter

        self.doublequote = doublequote
        self.escapechar = escapechar
        self.skipinitialspace = skipinitialspace
        self.quotechar = quotechar
        self.quoting = quoting

        if dialect is not None:
            self._process_dialect(dialect)

        # header / index-related
        self.header = header

        if index_col is not None:
            if not isinstance(index_col, (list, tuple, np.ndarray)):
                index_col = [index_col]
        self.index_col = index_col

        self.names = list(names) if names is not None else names

        # type conversion-related
        if converters is not None:
            assert(isinstance(converters, dict))
            self.converters = converters
        else:
            self.converters = {}

        self.thousands = thousands

        self.parse_dates = parse_dates
        self.keep_date_col = keep_date_col
        self.date_parser = date_parser
        self.dayfirst = dayfirst

        # Converting values to NA
        self.na_filter = na_filter
        self.na_values = self._get_na_values(na_values, keep_default_na)

        # skip rows, skip footer

        self.skip_footer = skip_footer

        if com.is_integer(skiprows):
            skiprows = range(skiprows)
        self.skiprows = set() if skiprows is None else set(skiprows)

        self.comment = comment

        # verbosity and error handling

        self.verbose = verbose
        self.warn_bad_lines = warn_bad_lines
        self.error_bad_lines = error_bad_lines

        # miscellanea
        self.kwds = kwds

        self.squeeze = squeeze    # return single-column DataFrame as Series

        self.as_recarray = as_recarray # C parser return structured array

        # file options
        self.f = f
        self.encoding = encoding
        self.engine_kind = engine

        self.factorize = factorize

        self.chunksize = chunksize

        self._engine = None
        self._make_engine(self.engine_kind)

    def _process_dialect(self, dialect):
        self.delimiter = dialect.delimiter
        self.doublequote = dialect.doublequote
        self.escapechar = dialect.escapechar
        self.skipinitialspace = dialect.skipinitialspace
        self.quotechar = dialect.quotechar
        self.quoting = dialect.quoting

    def __iter__(self):
        try:
            while True:
                yield self.read(self.chunksize)
        except StopIteration:
            pass

    def _get_na_values(self, na_values, keep_default_na=True):
        if na_values is None and keep_default_na:
            na_values = _NA_VALUES
        elif isinstance(na_values, dict):
            if keep_default_na:
                for k, v in na_values.iteritems():
                    v = set(list(v)) | _NA_VALUES
                    na_values[k] = v
        else:
            if not com.is_list_like(na_values):
                na_values = [na_values]
            na_values = set(list(na_values))
            if keep_default_na:
                na_values = na_values | _NA_VALUES

        return na_values

    def _make_engine(self, kind='c'):
        params = dict(delimiter=self.delimiter,
                      names=self.names,
                      index_col=self.index_col,
                      header=self.header,
                      skiprows=self.skiprows,
                      skip_footer=self.skip_footer,
                      thousands=self.thousands,
                      comment=self.comment,
                      doublequote=self.doublequote,
                      escapechar=self.escapechar,
                      skipinitialspace=self.skipinitialspace,
                      quotechar=self.quotechar,
                      quoting=self.quoting,
                      converters=self.converters,
                      parse_dates=self.parse_dates,
                      date_parser=self.date_parser,
                      keep_date_col=self.keep_date_col,
                      dayfirst=self.dayfirst,
                      na_values=self.na_values,
                      na_filter=self.na_filter,
                      verbose=self.verbose)

        params.update(self.kwds)

        if kind == 'c':
            params.update(warn_bad_lines=self.warn_bad_lines,
                          error_bad_lines=self.error_bad_lines,
                          as_recarray=self.as_recarray,
                          factorize=self.factorize)

            self._engine = CParserWrapper(self.f, **params)
        else:
            params.update(encoding=self.encoding)

            if kind == 'python':
                klass = PythonParser
            elif kind == 'python-fwf':

                klass = FixedWidthFieldParser

            self._engine = klass(self.f, **params)

    def _failover_to_python(self):
        raise NotImplementedError

    def read(self, nrows=None):
        if nrows is not None and self.skip_footer:
            raise ValueError('skip_footer not supported for iteration')

        # index = None

        ret = self._engine.read(nrows)
        index, columns, col_dict = ret

        # May alter columns / col_dict
        # index, columns, col_dict = self._create_index(col_dict, columns)

        df = DataFrame(col_dict, columns=columns, index=index)

        if self.squeeze and len(df.columns) == 1:
            return df[df.columns[0]]
        return df

    def _create_index(self, col_dict, columns):
        pass

    # backwards compatibility
    get_chunk = read


class ParserBase(object):

    def __init__(self, kwds):
        self.names = kwds.get('names')
        self.orig_names = None

        self.index_col = kwds.pop('index_col', None)
        self.index_names = None

        self.parse_dates = kwds.pop('parse_dates', False)
        self.date_parser = kwds.pop('date_parser', None)
        self.dayfirst = kwds.pop('dayfirst', False)
        self.keep_date_col = kwds.pop('keep_date_col', False)

        self.na_values = kwds.get('na_values')

        self._date_conv = _make_date_converter(date_parser=self.date_parser,
                                               dayfirst=self.dayfirst)

        self._name_processed = False

    @property
    def _has_complex_date_col(self):
        return (isinstance(self.parse_dates, dict) or
                (isinstance(self.parse_dates, list) and
                 len(self.parse_dates) > 0 and
                 isinstance(self.parse_dates[0], list)))

    def _should_parse_dates(self, i):
        if isinstance(self.parse_dates, bool):
            return self.parse_dates
        else:
            name = self.index_names[i]
            j = self.index_col[i]

            if np.isscalar(self.parse_dates):
                return (j == self.parse_dates) or (name == self.parse_dates)
            else:
                return (j in self.parse_dates) or (name in self.parse_dates)

    def _make_index(self, data, alldata, columns):
        if self.index_col is None:
            index = None

        elif not self._has_complex_date_col:
            index = self._get_simple_index(alldata, columns)
            index = self._agg_index(index)

        elif self._has_complex_date_col:
            if not self._name_processed:
                (self.index_names, _,
                 self.index_col) = _clean_index_names(list(columns),
                                                      self.index_col)
                self._name_processed = True
            index = self._get_complex_date_index(data, columns)
            index = self._agg_index(index, try_parse_dates=False)

        return index

    _implicit_index = False

    def _get_simple_index(self, data, columns):
        def ix(col):
            if not isinstance(col, basestring):
                return col
            raise ValueError('Index %s invalid' % col)
        index = None

        to_remove = []
        index = []
        for idx in self.index_col:
            i = ix(idx)
            to_remove.append(i)
            index.append(data[i])

        # remove index items from content and columns, don't pop in
        # loop
        for i in reversed(sorted(to_remove)):
            data.pop(i)
            if not self._implicit_index:
                columns.pop(i)

        return index

    def _get_complex_date_index(self, data, col_names):
        def _get_name(icol):
            if isinstance(icol, basestring):
                return icol

            if col_names is None:
                raise ValueError(('Must supply column order to use %s as '
                                  'index') % str(icol))

            for i, c in enumerate(col_names):
                if i == icol:
                    return c

        index = None

        to_remove = []
        index = []
        for idx in self.index_col:
            name = _get_name(idx)
            to_remove.append(name)
            index.append(data[name])

        # remove index items from content and columns, don't pop in
        # loop
        for c in reversed(sorted(to_remove)):
            data.pop(c)
            col_names.remove(c)

        return index

    def _agg_index(self, index, try_parse_dates=True):
        arrays = []
        for i, arr in enumerate(index):

            if (try_parse_dates and self._should_parse_dates(i)):
                arr = self._date_conv(arr)

            col_na_values = self.na_values

            if isinstance(self.na_values, dict):
                col_name = self.index_names[i]
                if col_name is not None:
                    col_na_values = _get_na_values(col_name,
                                                   self.na_values)

            arr, _ = _convert_types(arr, col_na_values)
            arrays.append(arr)

        index = MultiIndex.from_arrays(arrays, names=self.index_names)

        return index

    def _do_date_conversions(self, names, data):
        # returns data, columns
        if self.parse_dates is not None:
            data, names = _process_date_conversion(
                data, self._date_conv, self.parse_dates, self.index_col,
                self.index_names, names, keep_date_col=self.keep_date_col)

        return names, data

    def _exclude_implicit_index(self, alldata):

        if self._implicit_index:
            excl_indices = self.index_col

            data = {}
            offset = 0
            for i, col in enumerate(self.orig_names):
                while i + offset in excl_indices:
                    offset += 1
                data[col] = alldata[i + offset]
        else:
            data = dict((k, v) for k, v in izip(self.orig_names, alldata))

        return data


class CParserWrapper(ParserBase):
    """

    """

    def __init__(self, src, **kwds):
        self.kwds = kwds
        kwds = kwds.copy()

        self.as_recarray = kwds.get('as_recarray', False)

        ParserBase.__init__(self, kwds)

        self._reader = _parser.TextReader(src, **kwds)

        if self._reader.header is None:
            self.names = None
        else:
            self.names = list(self._reader.header)

        if self.names is None:
            self.names = ['X.%d' % (i + 1)
                          for i in range(self._reader.table_width)]

        self.orig_names = self.names

        if not self._has_complex_date_col:
            if self._reader.leading_cols == 0 and self.index_col is not None:
                self._name_processed = True
                (self.index_names, self.names,
                 self.index_col) = _clean_index_names(self.names,
                                                      self.index_col)

        self._implicit_index = self._reader.leading_cols > 0

    def read(self, nrows=None):
        if self.as_recarray:
            # what to do if there are leading columns?
            return self._reader.read(nrows)

        data = self._reader.read(nrows)
        names = self.names

        if self._reader.leading_cols:
            if self._has_complex_date_col:
                raise NotImplementedError('file structure not yet supported')

            # implicit index, no index names
            arrays = []

            for i in range(self._reader.leading_cols):
                if self.index_col is None:
                    values = data.pop(i)
                else:
                    values = data.pop(self.index_col[i])

                values = self._maybe_parse_dates(values, i,
                                                 try_parse_dates=True)
                arrays.append(values)

            index = MultiIndex.from_arrays(arrays)

            # rename dict keys
            data = sorted(data.items())
            data = dict((k, v) for k, (i, v) in zip(names, data))

            names, data = self._do_date_conversions(names, data)

        else:
            # rename dict keys
            data = sorted(data.items())

            # ugh, mutation
            names = list(self.orig_names)

            # columns as list
            alldata = [x[1] for x in data]

            data = dict((k, v) for k, (i, v) in zip(self.orig_names, data))

            names, data = self._do_date_conversions(names, data)
            index = self._make_index(data, alldata, names)


        return index, names, data

    def _get_index_names(self):
        names = list(self._reader.header)
        idx_names = None

        if self._reader.leading_cols == 0 and self.index_col is not None:
            (idx_names, names,
             self.index_col) = _clean_index_names(names, self.index_col)

        return names, idx_names

    def _maybe_parse_dates(self, values, index, try_parse_dates=True):
        if try_parse_dates and self._should_parse_dates(index):
            values = self._date_conv(values)
        return values



def TextParser(*args, **kwds):
    """
    Converts lists of lists/tuples into DataFrames with proper type inference
    and optional (e.g. string to datetime) conversion. Also enables iterating
    lazily over chunks of large files

    Parameters
    ----------
    data : file-like object or list
    delimiter : separator character to use
    dialect : str or csv.Dialect instance, default None
        Ignored if delimiter is longer than 1 character
    names : sequence, default
    header : int, default 0
        Row to use to parse column labels. Defaults to the first row. Prior
        rows will be discarded
    index_col : int or list, default None
        Column or columns to use as the (possibly hierarchical) index
    na_values : iterable, default None
        Custom NA values
    keep_default_na : bool, default True
    thousands : str, default None
        Thousands separator
    comment : str, default None
        Comment out remainder of line
    parse_dates : boolean, default False
    keep_date_col : boolean, default False
    date_parser : function, default None
    skiprows : list of integers
        Row numbers to skip
    skip_footer : int
        Number of line at bottom of file to skip
    encoding : string, default None
        Encoding to use for UTF when reading/writing (ex. 'utf-8')
    squeeze : boolean, default False
        returns Series if only one column
    """
    kwds['engine'] = 'python'
    return TextFileReader(*args, **kwds)


class PythonParser(ParserBase):

    def __init__(self, f, delimiter=None, dialect=None, names=None, header=0,
                 index_col=None,
                 na_values=None,
                 na_filter=True,
                 thousands=None,
                 quotechar='"',
                 escapechar=None,
                 doublequote=True,
                 skipinitialspace=False,
                 quoting=csv.QUOTE_MINIMAL,
                 comment=None, parse_dates=False, keep_date_col=False,
                 date_parser=None, dayfirst=False,
                 chunksize=None, skiprows=None, skip_footer=0, converters=None,
                 verbose=False, encoding=None, squeeze=False):
        """
        Workhorse function for processing nested list into DataFrame

        Should be replaced by np.genfromtxt eventually?
        """
        self.data = None
        self.buf = []
        self.pos = 0
        self.names = names
        self.header = header
        self.index_col = index_col
        self.chunksize = chunksize
        self.encoding = encoding

        self.parse_dates = parse_dates
        self.date_parser = date_parser
        self.dayfirst = dayfirst
        self._date_conv = _make_date_converter(date_parser=date_parser,
                                               dayfirst=dayfirst)
        self.keep_date_col = keep_date_col

        self.skiprows = skiprows

        self.skip_footer = skip_footer
        self.delimiter = delimiter

        self.quotechar = quotechar
        self.escapechar = escapechar
        self.doublequote = doublequote
        self.skipinitialspace = skipinitialspace
        self.quoting = quoting
        self.dialect = dialect

        self.verbose = verbose
        self.converters = converters

        self.na_values = na_values
        self.na_filter = na_filter

        if not na_filter:
            raise NotImplementedError

        self.squeeze = squeeze

        self.thousands = thousands
        self.comment = comment
        self._comment_lines = []

        if hasattr(f, 'readline'):
            self._make_reader(f)
        else:
            self.data = f
        self.columns = self._infer_columns()

        # get popped off for index
        self.orig_names = list(self.columns)

        # needs to be cleaned/refactored
        # multiple date column thing turning into a real spaghetti factory

        self.index_names = None
        self._name_processed = False
        if not self._has_complex_date_col:
            (self.index_names,
             self.orig_names, _) = self._get_index_name(self.columns)
            self._name_processed = True
        self._first_chunk = True

    def _make_reader(self, f):
        sep = self.delimiter

        if sep is None or len(sep) == 1:
            if self.dialect is not None:
                if isinstance(self.dialect, basestring):
                    dialect = csv.get_dialect(self.dialect)
                dia = dialect
            else:
                class MyDialect(csv.Dialect):
                    delimiter = self.delimiter
                    quotechar = self.quotechar
                    escapechar = self.escapechar
                    doublequote = self.doublequote
                    skipinitialspace = self.skipinitialspace
                    quoting = self.quoting
                    lineterminator = '\n'

                dia = MyDialect

            sniff_sep = True

            if sep is not None:
                sniff_sep = False
                dia.delimiter = sep
            # attempt to sniff the delimiter
            if sniff_sep:
                line = f.readline()
                while self.pos in self.skiprows:
                    self.pos += 1
                    line = f.readline()

                line = self._check_comments([line])[0]

                self.pos += 1
                sniffed = csv.Sniffer().sniff(line)
                dia.delimiter = sniffed.delimiter
                if self.encoding is not None:
                    self.buf.extend(list(
                        com.UnicodeReader(StringIO(line),
                                          dialect=dia,
                                          encoding=self.encoding)))
                else:
                    self.buf.extend(list(csv.reader(StringIO(line),
                                                    dialect=dia)))

            if self.encoding is not None:
                reader = com.UnicodeReader(f, dialect=dia,
                                           encoding=self.encoding)
            else:
                reader = csv.reader(f, dialect=dia)
        else:
            reader = (re.split(sep, line.strip()) for line in f)

        self.data = reader

    def read(self, rows=None):
        try:
            content = self._get_lines(rows)
        except StopIteration:
            if self._first_chunk:
                content = []
            else:
                raise

        # done with first read, next time raise StopIteration
        self._first_chunk = False

        columns = list(self.orig_names)
        if len(content) == 0: # pragma: no cover
            # DataFrame with the right metadata, even though it's length 0
            return _get_empty_meta(self.orig_names,
                                   self.index_col,
                                   self.index_names)

        alldata = self._rows_to_cols(content)
        data = self._exclude_implicit_index(alldata)
        data = self._convert_data(data)
        columns, data = self._do_date_conversions(self.columns, data)
        index = self._make_index(data, alldata, columns)

        return index, columns, data

    # legacy
    get_chunk = read

    def _convert_data(self, data):
        # apply converters
        converted = set()
        for col, f in self.converters.iteritems():
            if isinstance(col, int) and col not in self.orig_names:
                col = self.orig_names[col]
            data[col] = lib.map_infer(data[col], f)
            converted.add(col)

        # do type conversions
        result = {}
        for c, values in data.iteritems():
            if c in converted:
                result[c] = values

            col_na_values = _get_na_values(c, self.na_values)
            cvals, na_count = _convert_types(values, col_na_values)
            result[c] = cvals
            if self.verbose and na_count:
                print 'Filled %d NA values in column %s' % (na_count, str(c))

        return result

    def _infer_columns(self):
        names = self.names
        if names is not None:
            self.header = None

        if self.header is not None:
            if len(self.buf) > 0:
                line = self.buf[0]
            else:
                line = self._next_line()

            while self.pos <= self.header:
                line = self._next_line()

            columns = []
            for i, c in enumerate(line):
                if c == '':
                    columns.append('Unnamed: %d' % i)
                else:
                    columns.append(c)

            counts = {}
            for i, col in enumerate(columns):
                cur_count = counts.get(col, 0)
                if cur_count > 0:
                    columns[i] = '%s.%d' % (col, cur_count)
                counts[col] = cur_count + 1
            self._clear_buffer()
        else:
            if len(self.buf) > 0:
                line = self.buf[0]
            else:
                line = self._next_line()

            ncols = len(line)
            if not names:
                columns = ['X.%d' % (i + 1) for i in range(ncols)]
            else:
                columns = names

        return columns

    def _next_line(self):
        if isinstance(self.data, list):
            while self.pos in self.skiprows:
                self.pos += 1

            try:
                line = self.data[self.pos]
            except IndexError:
                raise StopIteration
        else:
            while self.pos in self.skiprows:
                next(self.data)
                self.pos += 1

            line = next(self.data)

        line = self._check_comments([line])[0]
        line = self._check_thousands([line])[0]

        self.pos += 1
        self.buf.append(line)

        return line

    def _check_comments(self, lines):
        if self.comment is None:
            return lines
        ret = []
        for l in lines:
            rl = []
            for x in l:
                if (not isinstance(x, basestring) or
                    self.comment not in x):
                    rl.append(x)
                else:
                    x = x[:x.find(self.comment)]
                    if len(x) > 0:
                        rl.append(x)
                    break
            ret.append(rl)
        return ret

    def _check_thousands(self, lines):
        if self.thousands is None:
            return lines
        nonnum = re.compile('[^-^0-9^%s^.]+' % self.thousands)
        ret = []
        for l in lines:
            rl = []
            for x in l:
                if (not isinstance(x, basestring) or
                    self.thousands not in x or
                    nonnum.search(x.strip())):
                    rl.append(x)
                else:
                    rl.append(x.replace(',', ''))
            ret.append(rl)
        return ret

    def _clear_buffer(self):
        self.buf = []

    _implicit_index = False

    def _get_index_name(self, columns):
        orig_names = list(columns)
        columns = list(columns)

        try:
            line = self._next_line()
        except StopIteration:
            line = None

        try:
            next_line = self._next_line()
        except StopIteration:
            next_line = None

        index_name = None

        # implicitly index_col=0 b/c 1 fewer column names
        implicit_first_cols = 0
        if line is not None:
            implicit_first_cols = len(line) - len(columns)
            if next_line is not None:
                if len(next_line) == len(line) + len(columns):
                    # column and index names on diff rows
                    implicit_first_cols = 0

                    self.index_col = range(len(line))
                    self.buf = self.buf[1:]

                    for c in reversed(line):
                        columns.insert(0, c)

                    return line, columns, orig_names

        if implicit_first_cols > 0:
            self._implicit_index = True
            if self.index_col is None:
                self.index_col = range(implicit_first_cols)
            index_name = None

        else:
            (index_name, columns,
             self.index_col) = _clean_index_names(columns, self.index_col)

        return index_name, orig_names, columns

    def _rows_to_cols(self, content):
        zipped_content = list(lib.to_object_array(content).T)

        col_len = len(self.orig_names)
        zip_len = len(zipped_content)

        if self._implicit_index:
            col_len += len(self.index_col)

        if col_len != zip_len:
            row_num = -1
            i = 0
            for (i, l) in enumerate(content):
                if len(l) != col_len:
                    break

            footers = 0
            if self.skip_footer:
                footers = self.skip_footer
                if footers > 0:
                    footers = footers - self.pos
            row_num = self.pos - (len(content) - i - footers)

            msg = ('Expecting %d columns, got %d in row %d' %
                   (col_len, zip_len, row_num))
            raise ValueError(msg)

        return zipped_content

    def _get_lines(self, rows=None):
        source = self.data
        lines = self.buf

        # already fetched some number
        if rows is not None:
            rows -= len(self.buf)

        if isinstance(source, list):
            if self.pos > len(source):
                raise StopIteration
            if rows is None:
                lines.extend(source[self.pos:])
                self.pos = len(source)
            else:
                lines.extend(source[self.pos:self.pos+rows])
                self.pos += rows
        else:
            new_rows = []
            try:
                if rows is not None:
                    for _ in xrange(rows):
                        new_rows.append(next(source))
                    lines.extend(new_rows)
                else:
                    rows = 0
                    while True:
                        try:
                            new_rows.append(next(source))
                            rows += 1
                        except csv.Error, inst:
                            if 'newline inside string' in inst.message:
                                row_num = str(self.pos + rows)
                                msg = ('EOF inside string starting with line '
                                       + row_num)
                                raise Exception(msg)
                            raise
            except StopIteration:
                lines.extend(new_rows)
                if len(lines) == 0:
                    raise
            self.pos += len(new_rows)

        self.buf = []

        if self.skip_footer:
            lines = lines[:self.skip_footer]

        lines = self._check_comments(lines)
        return self._check_thousands(lines)


def _make_date_converter(date_parser=None, dayfirst=False):
    def converter(*date_cols):
        if date_parser is None:
            return lib.try_parse_dates(_concat_date_cols(date_cols),
                                       dayfirst=dayfirst)
        else:
            try:
                return date_parser(*date_cols)
            except Exception:
                try:
                    return generic_parser(date_parser, *date_cols)
                except Exception:
                    return lib.try_parse_dates(_concat_date_cols(date_cols),
                                               parser=date_parser,
                                               dayfirst=dayfirst)

    return converter


def _process_date_conversion(data_dict, converter, parse_spec,
                             index_col, index_names, columns,
                             keep_date_col=False):
    def _isindex(colspec):
        return colspec in index_col or colspec in index_names

    new_cols = []
    new_data = {}

    orig_names = columns
    columns = list(columns)

    date_cols = set()

    if parse_spec is None or isinstance(parse_spec, bool):
        return data_dict, columns

    if isinstance(parse_spec, list):
        # list of column lists
        for colspec in parse_spec:
            if np.isscalar(colspec):
                if isinstance(colspec, int) and colspec not in data_dict:
                    colspec = orig_names[colspec]
                if _isindex(colspec):
                    continue
                data_dict[colspec] = converter(data_dict[colspec])
            else:
                new_name, col, old_names = _try_convert_dates(
                    converter, colspec, data_dict, orig_names)
                if new_name in data_dict:
                    raise ValueError('New date column already in dict %s' %
                                     new_name)
                new_data[new_name] = col
                new_cols.append(new_name)
                date_cols.update(old_names)

    elif isinstance(parse_spec, dict):
        # dict of new name to column list
        for new_name, colspec in parse_spec.iteritems():
            if new_name in data_dict:
                raise ValueError('Date column %s already in dict' %
                                 new_name)

            _, col, old_names = _try_convert_dates(converter, colspec,
                                                   data_dict, orig_names)

            new_data[new_name] = col
            new_cols.append(new_name)
            date_cols.update(old_names)

    data_dict.update(new_data)
    new_cols.extend(columns)

    if not keep_date_col:
        for c in list(date_cols):
            data_dict.pop(c)
            new_cols.remove(c)

    return data_dict, new_cols


def _try_convert_dates(parser, colspec, data_dict, columns):
    colset = set(columns)
    colnames = []

    for c in colspec:
        if c in colset:
            colnames.append(str(c))
        elif isinstance(c, int):
            colnames.append(str(columns[c]))

    new_name = '_'.join(colnames)
    to_parse = [data_dict[c] for c in colnames if c in data_dict]

    try:
        new_col = parser(*to_parse)
    except DateConversionError:
        new_col = parser(_concat_date_cols(to_parse))
    return new_name, new_col, colnames


def _clean_index_names(columns, index_col):
    if index_col is None:
        return None, columns, index_col

    columns = list(columns)

    cp_cols = list(columns)
    index_names = []

    # don't mutate
    index_col = list(index_col)

    for i, c in enumerate(index_col):
        if isinstance(c, basestring):
            index_names.append(c)
            for j, name in enumerate(cp_cols):
                if name == c:
                    index_col[i] = j
                    columns.remove(name)
                    break
        else:
            name = cp_cols[c]
            columns.remove(name)
            index_names.append(name)

    # hack
    if index_names[0] is not None and 'Unnamed' in index_names[0]:
        index_names[0] = None

    return index_names, columns, index_col


def _get_empty_meta(columns, index_col, index_names):
    columns = list(columns)

    if index_col is not None:
        index = MultiIndex.from_arrays([[]] * len(index_col),
                                       names=index_names)
        for n in index_col:
            columns.pop(n)
    else:
        index = Index([])

    return index, columns, {}


def _get_na_values(col, na_values):
    if isinstance(na_values, dict):
        if col in na_values:
            return set(list(na_values[col]))
        else:
            return _NA_VALUES
    else:
        return na_values

def _convert_to_ndarrays(dct, na_values, verbose=False):
    result = {}
    for c, values in dct.iteritems():
        col_na_values = _get_na_values(c, na_values)
        cvals, na_count = _convert_types(values, col_na_values)
        result[c] = cvals
        if verbose and na_count:
            print 'Filled %d NA values in column %s' % (na_count, str(c))
    return result

def _convert_types(values, na_values):
    na_count = 0
    if issubclass(values.dtype.type, (np.number, np.bool_)):
        mask = lib.ismember(values, na_values)
        na_count = mask.sum()
        if na_count > 0:
            if com.is_integer_dtype(values):
                values = values.astype(np.float64)
            np.putmask(values, mask, np.nan)
        return values, na_count

    try:
        result = lib.maybe_convert_numeric(values, na_values, False)
    except Exception:
        na_count = lib.sanitize_objects(values, na_values, False)
        result = values

    if result.dtype == np.object_:
        result = lib.maybe_convert_bool(values)

    return result, na_count


def _concat_date_cols(date_cols):
    if len(date_cols) == 1:
        return np.array([str(x) for x in date_cols[0]], dtype=object)

    # stripped = [map(str.strip, x) for x in date_cols]
    rs = np.array([' '.join([str(y) for y in x])
                   for x in zip(*date_cols)], dtype=object)
    return rs


class FixedWidthReader(object):
    """
    A reader of fixed-width lines.
    """
    def __init__(self, f, colspecs, filler, thousands=None):
        self.f = f
        self.colspecs = colspecs
        self.filler = filler # Empty characters between fields.
        self.thousands = thousands

        assert isinstance(colspecs, (tuple, list))
        for colspec in colspecs:
            assert isinstance(colspec, (tuple, list))
            assert len(colspec) == 2
            assert isinstance(colspec[0], int)
            assert isinstance(colspec[1], int)

    def next(self):
        line = next(self.f)
        # Note: 'colspecs' is a sequence of half-open intervals.
        return [line[fromm:to].strip(self.filler or ' ')
                for (fromm, to) in self.colspecs]

    # Iterator protocol in Python 3 uses __next__()
    __next__ = next


class FixedWidthFieldParser(PythonParser):
    """
    Specialization that Converts fixed-width fields into DataFrames.
    See PythonParser for details.
    """
    def __init__(self, f, **kwds):
        # Support iterators, convert to a list.
        self.colspecs = list(kwds.pop('colspecs'))

        PythonParser.__init__(self, f, **kwds)

    def _make_reader(self, f):
        self.data = FixedWidthReader(f, self.colspecs, self.delimiter)


#----------------------------------------------------------------------
# ExcelFile class

_openpyxl_msg = ("\nFor parsing .xlsx files 'openpyxl' is required.\n"
                 "You can install it via 'easy_install openpyxl' or "
                 "'pip install openpyxl'.\nAlternatively, you could save"
                 " the .xlsx file as a .xls file.\n")


class ExcelFile(object):
    """
    Class for parsing tabular excel sheets into DataFrame objects.
    Uses xlrd for parsing .xls files or openpyxl for .xlsx files.
    See ExcelFile.parse for more documentation

    Parameters
    ----------
    path : string or file-like object
        Path to xls file
    kind : {'xls', 'xlsx', None}, default None
    """
    def __init__(self, path_or_buf):
        self.use_xlsx = True
        self.path_or_buf = path_or_buf
        self.tmpfile = None

        if isinstance(path_or_buf, basestring):
            if path_or_buf.endswith('.xls'):
                self.use_xlsx = False
                import xlrd
                self.book = xlrd.open_workbook(path_or_buf)
            else:
                try:
                    from openpyxl.reader.excel import load_workbook
                    self.book = load_workbook(path_or_buf, use_iterators=True)
                except ImportError:  # pragma: no cover
                    raise ImportError(_openpyxl_msg)
        else:
            data = path_or_buf.read()

            try:
                import xlrd
                self.book = xlrd.open_workbook(file_contents=data)
                self.use_xlsx = False
            except Exception:
                from openpyxl.reader.excel import load_workbook
                buf = py3compat.BytesIO(data)
                self.book = load_workbook(buf, use_iterators=True)

    def __repr__(self):
        return object.__repr__(self)

    def parse(self, sheetname, header=0, skiprows=None, index_col=None,
              parse_cols=None, parse_dates=False, date_parser=None,
              na_values=None, thousands=None, chunksize=None):
        """
        Read Excel table into DataFrame

        Parameters
        ----------
        sheetname : string
            Name of Excel sheet
        header : int, default 0
            Row to use for the column labels of the parsed DataFrame
        skiprows : list-like
            Row numbers to skip (0-indexed)
        index_col : int, default None
            Column to use as the row labels of the DataFrame. Pass None if
            there is no such column
        parse_cols : int or list, default None
            If None then parse all columns,
            If int then indicates last column to be parsed
            If list of ints then indicates list of column numbers to be parsed
        na_values : list-like, default None
            List of additional strings to recognize as NA/NaN

        Returns
        -------
        parsed : DataFrame
        """
        choose = {True:self._parse_xlsx,
                  False:self._parse_xls}
        return choose[self.use_xlsx](sheetname, header=header,
                                     skiprows=skiprows, index_col=index_col,
                                     parse_cols=parse_cols,
                                     parse_dates=parse_dates,
                                     date_parser=date_parser,
                                     na_values=na_values,
                                     thousands=thousands,
                                     chunksize=chunksize)

    def _should_parse(self, i, parse_cols):
        if isinstance(parse_cols, int):
            return i <= parse_cols
        else:
            return i in parse_cols

    def _parse_xlsx(self, sheetname, header=0, skiprows=None, index_col=None,
                    parse_cols=None, parse_dates=False, date_parser=None,
                    na_values=None, thousands=None, chunksize=None):
        sheet = self.book.get_sheet_by_name(name=sheetname)
        data = []

        # it brings a new method: iter_rows()
        should_parse = {}

        for row in sheet.iter_rows():
            row_data = []
            for j, cell in enumerate(row):

                if parse_cols is not None and j not in should_parse:
                    should_parse[j] = self._should_parse(j, parse_cols)

                if parse_cols is None or should_parse[j]:
                    row_data.append(cell.internal_value)
            data.append(row_data)

        if header is not None:
            data[header] = _trim_excel_header(data[header])

        parser = TextParser(data, header=header, index_col=index_col,
                            na_values=na_values,
                            thousands=thousands,
                            parse_dates=parse_dates,
                            date_parser=date_parser,
                            skiprows=skiprows,
                            chunksize=chunksize)

        return parser.read()

    def _parse_xls(self, sheetname, header=0, skiprows=None, index_col=None,
                   parse_cols=None, parse_dates=False, date_parser=None,
                   na_values=None, thousands=None, chunksize=None):
        from datetime import MINYEAR, time, datetime
        from xlrd import xldate_as_tuple, XL_CELL_DATE, XL_CELL_ERROR

        datemode = self.book.datemode
        sheet = self.book.sheet_by_name(sheetname)

        data = []
        should_parse = {}
        for i in range(sheet.nrows):
            row = []
            for j, (value, typ) in enumerate(izip(sheet.row_values(i),
                                                  sheet.row_types(i))):
                if parse_cols is not None and j not in should_parse:
                    should_parse[j] = self._should_parse(j, parse_cols)

                if parse_cols is None or should_parse[j]:
                    if typ == XL_CELL_DATE:
                        dt = xldate_as_tuple(value, datemode)
                        # how to produce this first case?
                        if dt[0] < MINYEAR: # pragma: no cover
                            value = time(*dt[3:])
                        else:
                            value = datetime(*dt)
                    if typ == XL_CELL_ERROR:
                        value = np.nan
                    row.append(value)
            data.append(row)

        if header is not None:
            data[header] = _trim_excel_header(data[header])

        parser = TextParser(data, header=header, index_col=index_col,
                            na_values=na_values,
                            thousands=thousands,
                            parse_dates=parse_dates,
                            date_parser=date_parser,
                            skiprows=skiprows,
                            chunksize=chunksize)

        return parser.read()

    @property
    def sheet_names(self):
        if self.use_xlsx:
            return self.book.get_sheet_names()
        else:
            return self.book.sheet_names()


def _trim_excel_header(row):
    # trim header row so auto-index inference works
    while len(row) > 0 and row[0] == '':
        row = row[1:]
    return row

class ExcelWriter(object):
    """
    Class for writing DataFrame objects into excel sheets, uses xlwt for xls,
    openpyxl for xlsx.  See DataFrame.to_excel for typical usage.

    Parameters
    ----------
    path : string
        Path to xls file
    """
    def __init__(self, path):
        self.use_xlsx = True
        if path.endswith('.xls'):
            self.use_xlsx = False
            import xlwt
            self.book = xlwt.Workbook()
            self.fm_datetime = xlwt.easyxf(num_format_str='YYYY-MM-DD HH:MM:SS')
            self.fm_date = xlwt.easyxf(num_format_str='YYYY-MM-DD')
        else:
            from openpyxl.workbook import Workbook
            self.book = Workbook(optimized_write = True)
        self.path = path
        self.sheets = {}
        self.cur_sheet = None

    def save(self):
        """
        Save workbook to disk
        """
        self.book.save(self.path)

    def writerow(self, row, sheet_name=None):
        """
        Write the given row into Excel an excel sheet

        Parameters
        ----------
        row : list
            Row of data to save to Excel sheet
        sheet_name : string, default None
            Name of Excel sheet, if None, then use self.cur_sheet
        """
        if sheet_name is None:
            sheet_name = self.cur_sheet
        if sheet_name is None:  # pragma: no cover
            raise Exception('Must pass explicit sheet_name or set '
                            'cur_sheet property')
        if self.use_xlsx:
            self._writerow_xlsx(row, sheet_name)
        else:
            self._writerow_xls(row, sheet_name)

    def _writerow_xls(self, row, sheet_name):
        if sheet_name in self.sheets:
            sheet, row_idx = self.sheets[sheet_name]
        else:
            sheet = self.book.add_sheet(sheet_name)
            row_idx = 0
        sheetrow = sheet.row(row_idx)
        for i, val in enumerate(row):
            if isinstance(val, (datetime.datetime, datetime.date)):
                if isinstance(val, datetime.datetime):
                    sheetrow.write(i,val, self.fm_datetime)
                else:
                    sheetrow.write(i,val, self.fm_date)
            elif isinstance(val, np.int64):
                sheetrow.write(i,int(val))
            elif isinstance(val, np.bool8):
                sheetrow.write(i,bool(val))
            else:
                sheetrow.write(i,val)
        row_idx += 1
        if row_idx == 1000:
            sheet.flush_row_data()
        self.sheets[sheet_name] = (sheet, row_idx)

    def _writerow_xlsx(self, row, sheet_name):
        if sheet_name in self.sheets:
            sheet, row_idx = self.sheets[sheet_name]
        else:
            sheet = self.book.create_sheet()
            sheet.title = sheet_name
            row_idx = 0

        conv_row = []
        for val in row:
            if isinstance(val, np.int64):
                val = int(val)
            elif isinstance(val, np.bool8):
                val = bool(val)
            conv_row.append(val)
        sheet.append(conv_row)
        row_idx += 1
        self.sheets[sheet_name] = (sheet, row_idx)
